import argparse
import json
import re
from collections import Counter
from datetime import date, datetime
from pathlib import Path

import openpyxl


def clean(value):
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value).strip()


def phone(value):
    text = clean(value)
    if not text:
        return ""
    return re.sub(r"\.0$", "", text)


def compact_notes(row):
    notes = []
    for i in range(1, 11):
        value = clean(row.get(f"會員備註{i}"))
        if value:
            notes.append(value)
    return notes


def gender_label(value):
    raw = clean(value).lower()
    if raw in {"man", "male", "m"}:
        return "男"
    if raw in {"woman", "female", "f"}:
        return "女"
    return clean(value)


def to_member(row, imported_at):
    legacy_id = clean(row.get("記錄編號"))
    mobile = phone(row.get("行動電話"))
    tel = phone(row.get("室內電話"))
    status = clean(row.get("會員狀態")) or "正常"
    city = clean(row.get("縣市"))
    district = clean(row.get("地區"))
    address = clean(row.get("通訊地址"))
    zip_code = clean(row.get("郵遞區號"))
    notes = compact_notes(row)
    registered_at = clean(row.get("註冊日期"))
    updated_at = clean(row.get("更新日期")) or imported_at
    member = {
        "userId": legacy_id,
        "legacyMemberId": legacy_id,
        "name": clean(row.get("會員名稱")) or legacy_id,
        "displayName": clean(row.get("會員名稱")) or legacy_id,
        "phone": mobile or tel,
        "mobile": mobile,
        "tel": tel,
        "zipCode": zip_code,
        "city": city,
        "district": district,
        "address": address,
        "fullAddress": "".join([zip_code, city, district, address]),
        "gender": gender_label(row.get("性別")),
        "legacyGender": clean(row.get("性別")),
        "status": status,
        "memberStatus": status,
        "memberTier": "一般會員",
        "edmSubscribed": "未訂閱" not in clean(row.get("訂閱eDM")),
        "edmStatus": clean(row.get("訂閱eDM")),
        "subscriptionToken": clean(row.get("訂閱憑證")),
        "referrerId": clean(row.get("推薦會員")),
        "birthday": clean(row.get("生日")),
        "registeredAt": registered_at,
        "createdAt": registered_at or imported_at,
        "updatedAt": updated_at,
        "lastLoginAt": clean(row.get("上次登入")),
        "notes": notes,
        "note": "；".join(notes),
        "source": "legacy_member_xlsx",
        "sourceFile": "Member_list_2026_05_31.xlsx",
        "importedAt": imported_at,
    }
    if status and status != "正常":
        member["isDeleted"] = True
    return member


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", required=True)
    parser.add_argument("--out", default="tools/hooktea-members-kv-bulk.json")
    parser.add_argument("--report", default="tools/hooktea-members-import-report.json")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    headers = [clean(x) for x in next(rows_iter)]
    imported_at = datetime.utcnow().replace(microsecond=0).isoformat() + "Z"

    bulk = []
    seen = set()
    duplicate_ids = []
    status_counts = Counter()
    missing_phone = 0
    for values in rows_iter:
        row = {headers[i]: values[i] if i < len(values) else "" for i in range(len(headers))}
        legacy_id = clean(row.get("記錄編號"))
        if not legacy_id:
            continue
        if legacy_id in seen:
            duplicate_ids.append(legacy_id)
            continue
        seen.add(legacy_id)
        member = to_member(row, imported_at)
        status_counts[member["memberStatus"]] += 1
        if not member["phone"]:
            missing_phone += 1
        bulk.append({
            "key": f"USER_{legacy_id}",
            "value": json.dumps(member, ensure_ascii=False, separators=(",", ":")),
        })

    Path(args.out).parent.mkdir(parents=True, exist_ok=True)
    Path(args.out).write_text(json.dumps(bulk, ensure_ascii=False, indent=2), encoding="utf-8")
    report = {
        "source": str(xlsx_path),
        "sheet": ws.title,
        "headers": headers,
        "importedAt": imported_at,
        "totalPrepared": len(bulk),
        "duplicateIdsSkipped": duplicate_ids,
        "statusCounts": dict(status_counts),
        "missingPhone": missing_phone,
        "output": args.out,
    }
    Path(args.report).write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
