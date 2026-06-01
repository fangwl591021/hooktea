import argparse
import json
import re
import sys
from collections import Counter, defaultdict
from datetime import date, datetime, timezone
from pathlib import Path

import openpyxl


def clean(value):
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat(sep=" ")
    return str(value).strip()


def normalize_phone(value):
    text = clean(value)
    text = re.sub(r"\.0$", "", text)
    digits = re.sub(r"\D+", "", text)
    if digits and not digits.startswith("0") and len(digits) == 9:
        digits = "0" + digits
    return digits


def parse_points(value):
    text = clean(value).replace(",", "")
    if not text:
        return 0
    return int(float(text))


def load_members(path):
    items = json.loads(Path(path).read_text(encoding="utf-8"))
    members = []
    for item in items:
        data = json.loads(item["value"])
        members.append(data)
    return members


def build_member_indexes(members):
    by_phone = defaultdict(dict)
    by_name = defaultdict(dict)
    for member in members:
        uid = clean(member.get("userId"))
        if not uid:
            continue
        for key in ["phone", "mobile", "tel"]:
            phone = normalize_phone(member.get(key))
            if phone:
                by_phone[phone][uid] = member
        name = clean(member.get("name") or member.get("displayName"))
        if name:
            by_name[name][uid] = member
    return {k: list(v.values()) for k, v in by_phone.items()}, {k: list(v.values()) for k, v in by_name.items()}


def choose_member(row, by_phone, by_name):
    phone = normalize_phone(row.get("會員電話"))
    name = clean(row.get("會員名稱"))
    if phone:
        matches = by_phone.get(phone, [])
        if len(matches) == 1:
            return matches[0], "phone"
        if len(matches) > 1 and name:
            named = [m for m in matches if clean(m.get("name") or m.get("displayName")) == name]
            if len(named) == 1:
                return named[0], "phone_name"
            return None, "duplicate_phone"
    if name:
        matches = by_name.get(name, [])
        if len(matches) == 1:
            return matches[0], "name"
        if len(matches) > 1:
            return None, "duplicate_name"
    return None, "unmatched"


def main():
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", required=True)
    parser.add_argument("--members", default="tools/hooktea-members-kv-bulk.json")
    parser.add_argument("--out", default="tools/hooktea-points-kv-bulk.json")
    parser.add_argument("--report", default="tools/hooktea-points-import-report.json")
    args = parser.parse_args()

    members = load_members(args.members)
    by_phone, by_name = build_member_indexes(members)
    wb = openpyxl.load_workbook(args.xlsx, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    headers = [clean(x) for x in next(rows_iter)]
    imported_at = datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")

    totals = Counter()
    point_totals = defaultdict(int)
    imported_rows = defaultdict(list)
    skipped = []

    for values in rows_iter:
        row = {headers[i]: values[i] if i < len(values) else "" for i in range(len(headers))}
        name = clean(row.get("會員名稱"))
        amount = parse_points(row.get("購物金/優惠券"))
        totals["rows"] += 1
        if amount == 0:
            totals["zeroSkipped"] += 1
            continue
        if amount < 0:
            totals["negativeSkipped"] += 1
            continue
        member, match_type = choose_member(row, by_phone, by_name)
        if not member:
            totals[f"{match_type}Skipped"] += 1
            skipped.append({
                "name": name,
                "phone": clean(row.get("會員電話")),
                "points": amount,
                "reason": match_type,
            })
            continue
        uid = member["userId"]
        point_totals[uid] += amount
        imported_rows[uid].append({
            "name": name,
            "phone": clean(row.get("會員電話")),
            "points": amount,
            "sourceUpdatedAt": clean(row.get("更新日期")),
            "matchType": match_type,
        })
        totals["positiveMatched"] += 1

    bulk = []
    for uid, balance in sorted(point_totals.items()):
        rows = imported_rows[uid]
        member_name = clean(rows[0].get("name"))
        logs = [{
            "logId": f"legacy_coupon_{uid}",
            "amount": balance,
            "reason": "舊系統購物金/優惠券轉入",
            "createdAt": imported_at,
            "type": "EARN",
            "source": "legacy_coupon_xlsx",
            "sourceRows": rows,
        }]
        data = {
            "balance": balance,
            "logs": logs,
            "legacyCouponBalance": balance,
            "source": "legacy_coupon_xlsx",
            "sourceFile": "Coupon_list_2026_05_31 (1).xlsx",
            "memberName": member_name,
            "importedAt": imported_at,
        }
        bulk.append({
            "key": f"POINTS_{uid}",
            "value": json.dumps(data, ensure_ascii=False, separators=(",", ":")),
        })

    Path(args.out).parent.mkdir(parents=True, exist_ok=True)
    Path(args.out).write_text(json.dumps(bulk, ensure_ascii=False, indent=2), encoding="utf-8")
    report = {
        "source": args.xlsx,
        "sheet": ws.title,
        "headers": headers,
        "importedAt": imported_at,
        "totalRows": totals["rows"],
        "positiveRowsMatched": totals["positiveMatched"],
        "membersPrepared": len(bulk),
        "totalPositivePoints": sum(point_totals.values()),
        "zeroSkipped": totals["zeroSkipped"],
        "negativeSkipped": totals["negativeSkipped"],
        "unmatchedPositiveSkipped": totals["unmatchedSkipped"],
        "duplicatePhoneSkipped": totals["duplicate_phoneSkipped"],
        "duplicateNameSkipped": totals["duplicate_nameSkipped"],
        "skippedPositiveRows": skipped,
        "output": args.out,
    }
    Path(args.report).write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
